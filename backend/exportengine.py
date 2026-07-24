#exportengine.py
# All three export formats (xlsx/csv/json) go through here now, so every
# format agrees on the same flatten_invoice/flatten_products field set
# instead of Node hand-duplicating a separate (and incomplete - no product
# rows at all) flattening for csv. Format is inferred from output_path's
# extension. Invocation contract mirrors engine.py's runPython pattern
# (src/engine.js): argv = [payload_path, output_path], prints a single JSON
# object to stdout.

import sys
import json
import os
import csv
import hashlib
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINK_FONT = Font(color="0563C1", underline="single")

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALT_FILL = PatternFill("solid", fgColor="EEF2F7")


def style_header_row(ws, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def to_num(v):
    if v is None or v == "":
        return ""
    s = str(v).replace(",", "").replace("%", "").strip()
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except Exception:
        return v


def style_data_rows(ws, num_rows, num_cols):
    for r in range(2, num_rows + 2):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = BORDER
            if fill:
                cell.fill = fill


def autofit_columns(ws, rows_data, headers):
    for i, header in enumerate(headers, 1):
        max_len = len(header)
        for row in rows_data:
            val = str(row.get(header, "") or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 50)


def flatten_invoice(inv):
    p = inv.get("provider", {}) or {}
    r = inv.get("receiver", {}) or {}
    i = inv.get("invoice", {}) or {}
    t = inv.get("totals", {}) or {}
    invoice_id = i.get("invoice_id", "")
    _hash_src = f"{inv.get('batch','')}-{invoice_id}-{p.get('name','')}-{i.get('date_of_invoice','')}"
    product_list_id = hashlib.sha256(_hash_src.encode()).hexdigest()[:16].upper()

    return {
        "product_list_id": product_list_id,
        "provider_name": p.get("name", ""),
        "provider_address": p.get("address", ""),
        "provider_city": p.get("city", ""),
        "provider_state": p.get("state", ""),
        "provider_country": p.get("country", ""),
        "provider_pincode": p.get("pincode", ""),
        "provider_contact_no": p.get("contact_no", ""),
        "provider_pan_no": p.get("pan_no", ""),
        "provider_gstin": p.get("gstin", ""),
        "provider_cin_no": p.get("cin_no", ""),
        "provider_tan": p.get("tan", ""),
        "provider_email": p.get("email", ""),
        "provider_bank_name": p.get("bank_name", ""),
        "provider_account_no": p.get("account_no", ""),
        "provider_ifsc_code": p.get("ifsc_code", ""),
        "receiver_name": r.get("name", ""),
        "receiver_address": r.get("address", ""),
        "receiver_city": r.get("city", ""),
        "receiver_state": r.get("state", ""),
        "receiver_country": r.get("country", ""),
        "receiver_pincode": r.get("pincode", ""),
        "receiver_contact_no": r.get("contact_no", ""),
        "receiver_pan_no": r.get("pan_no", ""),
        "receiver_gstin": r.get("gstin", ""),
        "receiver_cin_no": r.get("cin_no", ""),
        "receiver_tan": r.get("tan", ""),
        "receiver_email": r.get("email", ""),
        "invoice_id": invoice_id,
        "date_of_invoice": i.get("date_of_invoice", ""),
        "place_of_supply": i.get("place_of_supply", ""),
        "vendor_code": i.get("vendor_code", ""),
        "due_date": i.get("due_date", ""),
        "order_number": i.get("order_number", ""),
        "transporter": i.get("transporter", ""),
        "overall_cgst_rate": to_num(t.get("overall_cgst_rate", "")),
        "overall_cgst": to_num(t.get("overall_cgst", "")),
        "overall_sgst_rate": to_num(t.get("overall_sgst_rate", "")),
        "overall_sgst": to_num(t.get("overall_sgst", "")),
        "overall_total_amount": to_num(t.get("overall_total_amount", "")),
    }


def flatten_products(inv):
    i = inv.get("invoice", {}) or {}
    invoice_id = i.get("invoice_id", "")
    p_info = inv.get("provider", {}) or {}
    _hash_src = f"{inv.get('batch','')}-{invoice_id}-{p_info.get('name','')}-{i.get('date_of_invoice','')}"
    product_list_id = hashlib.sha256(_hash_src.encode()).hexdigest()[:16].upper()
    rows = []
    for p in inv.get("products", []) or []:
        rows.append({
            "product_list_id": product_list_id,
            "invoice_id": invoice_id,
            "item_no": p.get("item_no", ""),
            "item_code": p.get("item_code", ""),
            "description": p.get("description", ""),
            "hsn_sac": p.get("hsn_sac", ""),
            "qty": to_num(p.get("qty", "")),
            "rate": to_num(p.get("rate", "")),
            "total_base_value": to_num(p.get("total_base_value", "")),
            "cgst_rate": to_num(p.get("cgst_rate", "")),
            "cgst_value": to_num(p.get("cgst_value", "")),
            "sgst_rate": to_num(p.get("sgst_rate", "")),
            "sgst_value": to_num(p.get("sgst_value", "")),
            "total_gst": p.get("total_gst", ""),
            "total_amount": p.get("total_amount", ""),
        })
    return rows


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_header_row(ws, len(headers))
    style_data_rows(ws, len(rows), len(headers))
    autofit_columns(ws, rows, headers)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def add_table(ws, headers, num_rows, table_name):
    if num_rows == 0:
        return
    ref = f"A1:{get_column_letter(len(headers))}{num_rows + 1}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def add_count_per_day_chart(wb, invoice_rows):
    ws = wb.create_sheet("Count Per Day")
    count_map = defaultdict(int)
    for row in invoice_rows:
        date = row.get("date_of_invoice") or "Unknown"
        count_map[date] += 1

    ws.append(["Date", "Invoice Count"])
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    ws.cell(1, 2).fill = HEADER_FILL

    for date, count in sorted(count_map.items()):
        ws.append([date, count])

    if len(count_map) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Invoices Per Day"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Date"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(count_map) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(count_map) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D2")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def export(payload, output_path):
    # payload: { invoices: [...already field-filtered invoice objects...] }
    invoices = payload.get("invoices", [])
    ext = os.path.splitext(output_path)[1].lower()

    if ext == ".json":
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump({"invoices": invoices}, f, indent=2)
        return output_path

    invoice_rows = [flatten_invoice(inv) for inv in invoices]

    if ext == ".csv":
        headers = list(invoice_rows[0].keys()) if invoice_rows else list(flatten_invoice({}).keys())
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(invoice_rows)
        return output_path

    product_rows = [p for inv in invoices for p in flatten_products(inv)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Invoices Sheet ---
    ws_inv = wb.create_sheet("Invoices")
    inv_headers = list(invoice_rows[0].keys()) if invoice_rows else list(flatten_invoice({}).keys())
    write_sheet(ws_inv, inv_headers, invoice_rows)
    add_table(ws_inv, inv_headers, len(invoice_rows), "InvoicesTable")

    if "product_list_id" in inv_headers:
        iid_col = inv_headers.index("product_list_id") + 1
        ws_inv.cell(row=1, column=iid_col).value = "IID"

    # --- Products Sheet ---
    ws_prod = wb.create_sheet("Products")
    prod_headers = list(product_rows[0].keys()) if product_rows else (
        list(flatten_products({"invoice": {}, "products": [{}]})[0].keys())
        if flatten_products({"invoice": {}, "products": [{}]}) else []
    )
    if prod_headers:
        write_sheet(ws_prod, prod_headers, product_rows)
        add_table(ws_prod, prod_headers, len(product_rows), "ProductsTable")
        if "product_list_id" in prod_headers:
            ws_prod.cell(row=1, column=prod_headers.index("product_list_id") + 1).value = "IID"

    # Hyperlink IID in Invoices sheet -> matching row in Products sheet
    pl_col = inv_headers.index("product_list_id") + 1 if "product_list_id" in inv_headers else None
    if pl_col:
        iid_to_row = {}
        for pr in range(2, len(product_rows) + 2):
            iid_val = ws_prod.cell(row=pr, column=1).value
            if iid_val and iid_val not in iid_to_row:
                iid_to_row[iid_val] = pr

        for r in range(2, len(invoice_rows) + 2):
            cell = ws_inv.cell(row=r, column=pl_col)
            iid_val = cell.value or ""
            target_row = iid_to_row.get(iid_val, 1)
            cell.hyperlink = f"#Products!A{target_row}"
            cell.font = LINK_FONT

    # --- Chart Sheet ---
    add_count_per_day_chart(wb, invoice_rows)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(json.dumps({"error": "Usage: exportengine.py <json_payload_path> <output_path>"}))
        sys.exit(1)

    payload_path = sys.argv[1]
    output_path = sys.argv[2]

    try:
        with open(payload_path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        result_path = export(payload, output_path)
        print(json.dumps({"success": True, "path": result_path}))
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
